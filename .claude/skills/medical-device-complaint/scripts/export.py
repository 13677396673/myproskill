#!/usr/bin/env python3
"""
医疗器械客诉单 - 多格式导出工具
Medical Device Complaint Form - Export Tool

用法:
  python scripts/export.py --input data.json --output report --format xlsx
  python scripts/export.py --input data.json --output report --format csv
  python scripts/export.py --input data.json --output report --format all
"""

import argparse
import csv
import json
import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from model import ComplaintForm

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"


# ── 展平数据 ──────────────────────────────────────────

def flatten_form(form: ComplaintForm) -> dict:
    """将 ComplaintForm 展平为单行字典（含 RPN 等派生字段）。"""
    d = form.to_dict()
    d["rpn"] = form.rpn if form.rpn is not None else ""
    d["risk_level"] = form.risk_level
    d["is_high_risk"] = "是" if form.is_high_risk else "否"
    return d


FIELD_LABELS_ZH = {
    "complaint_id": "客诉编号",
    "received_date": "接收日期",
    "source": "投诉来源",
    "reporter_name": "报告人",
    "hospital": "医院/机构",
    "department": "科室",
    "contact": "联系方式",
    "report_date": "报告日期",
    "product_name": "产品名称",
    "model": "型号规格",
    "lot_number": "批号/序列号",
    "manufacture_date": "生产日期",
    "expiry_date": "有效期",
    "category": "投诉类别",
    "description": "投诉详细描述",
    "patient_harm": "涉及患者伤害",
    "frequency": "发生频率",
    "investigation_date": "调查日期",
    "investigator": "调查人",
    "investigation_method": "调查方法",
    "investigation_findings": "调查发现",
    "root_cause": "根本原因分析",
    "root_cause_category": "根本原因分类",
    "immediate_action": "紧急处理措施",
    "capa_action": "长期纠正预防措施",
    "responsible_person": "责任人",
    "completion_date": "计划完成日期",
    "severity": "严重性 (S)",
    "occurrence": "发生概率 (O)",
    "detection": "可检测性 (D)",
    "rpn": "RPN",
    "risk_level": "风险等级",
    "is_high_risk": "高风险",
    "regulatory_reportable": "需要报告监管机构",
    "regulatory_status": "报告状态",
    "investigator_signature": "调查人签名",
    "investigator_sign_date": "调查人签名日期",
    "reviewer_name": "审核人",
    "reviewer_signature": "审核人签名",
    "reviewer_sign_date": "审核人签名日期",
    "approver_name": "批准人",
    "approver_signature": "批准人签名",
    "approver_sign_date": "批准人签名日期",
    "version": "版本号",
    "language": "语言",
}

FIELD_ORDER = [
    "complaint_id", "version", "received_date", "source",
    "reporter_name", "hospital", "department", "contact", "report_date",
    "product_name", "model", "lot_number", "manufacture_date", "expiry_date",
    "category", "description", "patient_harm", "frequency",
    "investigation_date", "investigator", "investigation_method",
    "investigation_findings", "root_cause", "root_cause_category",
    "immediate_action", "capa_action", "responsible_person", "completion_date",
    "severity", "occurrence", "detection", "rpn", "risk_level", "is_high_risk",
    "regulatory_reportable", "regulatory_status",
    "investigator_signature", "investigator_sign_date",
    "reviewer_name", "reviewer_signature", "reviewer_sign_date",
    "approver_name", "approver_signature", "approver_sign_date",
    "language",
]


# ── CSV 导出 ──────────────────────────────────────────

def export_csv(data: dict, output_path: str):
    """导出为 CSV（UTF-8 BOM 编码，Excel 可直接打开中文）。"""
    flat = flatten_form(ComplaintForm.from_dict(data))
    fieldnames = [f for f in FIELD_ORDER if f in flat]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writerow({k: FIELD_LABELS_ZH.get(k, k) for k in fieldnames})
        writer.writerow({k: flat.get(k, "") for k in fieldnames})

    print(f"CSV 已导出: {output_path}")


# ── Excel 导出 ────────────────────────────────────────

def export_xlsx(data: dict, output_path: str):
    """导出为 Excel（含条件格式：高风险行红色高亮）。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误: 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    form = ComplaintForm.from_dict(data)
    flat = flatten_form(form)
    wb = openpyxl.Workbook()

    # ── Sheet 1: 客诉单详情 ──
    ws = wb.active
    ws.title = "客诉单详情"

    fieldnames = [f for f in FIELD_ORDER if f in flat]
    header_font = Font(name="微软雅黑", bold=True, size=10)
    header_fill = PatternFill(start_color="E6F2FA", end_color="E6F2FA", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="B4BCC8"),
        right=Side(style="thin", color="B4BCC8"),
        top=Side(style="thin", color="B4BCC8"),
        bottom=Side(style="thin", color="B4BCC8"),
    )

    # 写表头（A 列 = 字段名，B 列 = 值）
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    for row_idx, key in enumerate(fieldnames, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=FIELD_LABELS_ZH.get(key, key))
        cell_a.font = header_font
        cell_a.fill = header_fill
        cell_a.alignment = header_alignment
        cell_a.border = thin_border

        cell_b = ws.cell(row=row_idx, column=2, value=str(flat.get(key, "")))
        cell_b.font = Font(name="微软雅黑", size=10)
        cell_b.border = thin_border

    # 高风险行高亮
    if form.is_high_risk:
        red_fill = PatternFill(start_color="FFEBEB", end_color="FFEBEB", fill_type="solid")
        for row_idx in range(1, len(fieldnames) + 1):
            ws.cell(row=row_idx, column=1).fill = red_fill
            ws.cell(row=row_idx, column=2).fill = red_fill

    # ── Sheet 2: 风险矩阵 ──
    ws2 = wb.create_sheet("风险矩阵")
    ws2.column_dimensions["A"].width = 14

    S_LABELS = ["严重性 S →"]
    O_LABELS = ["概率 O ↓"]
    for i in range(1, 6):
        ws2.column_dimensions[get_column_letter(i + 1)].width = 10
        S_LABELS.append(i)
    for i, label in enumerate(O_LABELS):
        ws2.cell(row=1, column=1 + i, value=label).font = Font(bold=True, size=10)

    for ox in range(5):
        row = ox + 2
        ws2.cell(row=row, column=1, value=5 - ox).font = Font(bold=True, size=10)
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for sx in range(5):
            col = sx + 2
            risk = (sx + 1) * (5 - ox)
            cell = ws2.cell(row=row, column=col, value=risk)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if risk <= 4:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif risk <= 9:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif risk <= 16:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")

            cell.border = thin_border

        # 标记当前 S,O 位置
        if form.severity is not None and form.occurrence is not None:
            sx_mark = form.severity - 1
            ox_mark = 4 - (form.occurrence - 1)
            mark_cell = ws2.cell(row=ox_mark + 2, column=sx_mark + 2)
            mark_cell.font = Font(bold=True, color="FF0000", size=12)
            mark_cell.value = f"*{mark_cell.value}"

    # UDF 表格实际超出范围
    wb.save(output_path)
    print(f"Excel 已导出: {output_path}")


# ── 主入口 ────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="医疗器械客诉单导出工具")
    parser.add_argument("--input", "-i", required=True, help="输入 JSON 文件")
    parser.add_argument("--output", "-o", default="", help="输出文件路径（不含扩展名）")
    parser.add_argument("--format", "-f", default="all",
                        choices=["xlsx", "csv", "all"], help="导出格式")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    output_stem = args.output or Path(args.input).stem

    if args.format in ("csv", "all"):
        csv_path = f"{output_stem}.csv"
        export_csv(data, csv_path)

    if args.format in ("xlsx", "all"):
        xlsx_path = f"{output_stem}.xlsx"
        export_xlsx(data, xlsx_path)

    print("导出完成。")


if __name__ == "__main__":
    main()
